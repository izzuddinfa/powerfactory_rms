import sys
sys.path.append(r"C:\Program Files\DIgSILENT\PowerFactory 2023 SP3A\Python\3.11")
import powerfactory as pf

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import gc
from openpyxl import Workbook
from time import perf_counter
from os import listdir
from os.path import isfile, join
import colormaps as cmaps 

class PowerFactorySim:
    def __init__(self, folder_name='', project_name='Project', study_case_name='Study Case'):
        # Aktifkan project
        self.folder_name = folder_name
        self.project_name = project_name
        self.study_case_name = study_case_name
        self.app = pf.GetApplication()
        self.project = self._activate_project()

        # Load komponen sistem tenaga listrik
        self.genObj = self.app.GetCalcRelevantObjects("*.ElmSym")
        self.termObj = self.app.GetCalcRelevantObjects("*.ElmTerm")
        self.lineObj = self.app.GetCalcRelevantObjects("*.ElmLne")
        self.trfObj = self.app.GetCalcRelevantObjects("*.ElmTr2")
        self.loadObj=self.app.GetCalcRelevantObjects("*.ElmLod")

    # Fungsi untuk mengaktifkan project
    def _activate_project(self):
        project_path = join(self.folder_name, self.project_name)
        project = self.app.ActivateProject(project_path)
        study_case_folder = self.app.GetProjectFolder('study')
        study_case = study_case_folder.GetContents(f'{self.study_case_name}.IntCase')[0]
        study_case.Activate()
        return project
    
    # Fungsi untuk menyimpan nilai default load dan generator
    def loadGenData(self):
        load = pd.DataFrame({
            "name": [obj.loc_name for obj in self.loadObj],
            "P": [obj.plini for obj in self.loadObj],
            "Q": [obj.qlini for obj in self.loadObj]
        })
        gen = pd.DataFrame({
            "name": [obj.loc_name for obj in self.genObj],
            "P": [obj.pgini for obj in self.genObj],
            "iv_mode": [obj.iv_mode for obj in self.genObj],
            "bus_typ": [obj.GetAttribute("e:bustp") for obj in self.genObj]
        })
        return load, gen

    # Fungsi untuk mengatur daya aktif dan daya reaktif beban
    def loadSetup(self, load_level, load_default):
        load_data = load_default.set_index("name")[["P", "Q"]].to_dict("index")
        for load in self.loadObj:
            if load.loc_name in load_data:
                load.plini = load_data[load.loc_name]["P"] * load_level
                load.qlini = load_data[load.loc_name]["Q"] * load_level
    
    # Fungsi untuk mengatur parameter simulasi Optimal Power Flow
    def opfSetup(self):
        for gen in self.genObj:
            gen.ictpg = 1 # Atur Control Active Power ke True
            gen.ictqg = 1 # Atur Control Reactive Power ke True
        for term in self.termObj:
            term.iOPFCvmax = 1 # Atur untuk menggunakan batas maks. tegangan bus ke True
            term.iOPFCvmin = 1 # Atur untuk menggunakan batas min. tegangan bus ke True
            term.vmin = 0.9 # Batas min. tegangan bus
            term.vmax = 1.05 # Batas maks. tegangan bus
        for line in self.lineObj:
            line.iOPFCload = 1 # Atur Max Loading Contraint menggunakan Hard Constraint
            line.maxload = 100 # Atur batas maks. loading saluran
        for trf in self.trfObj:
            trf.i_uopt = 1 # Atur tap changer ke On
            trf.ionlyPre = 1 # Atur dalam DC OPF berlaku untuk pre- dan post-fault position
            trf.i_uoptCont = 0 # Atur control mode tap changer ke continous

    # Fungsi untuk simulasi aliran daya
    def ldfAnalysis(self):
        self.ldf = self.app.GetFromStudyCase('ComLdf')
        self.ldf.iopt_plim = 1
        self.ldf.iopt_lim = 1
        self.ldf.Execute()

    # Fungsi untuk simulasi optimal power flow
    def opfAnalysis(self, iopt_obj='dev'):
        # Configure OPF settings
        self.opf = self.app.GetFromStudyCase("ComOpf")
        self.opf.iopt_ACDC = 0  # For AC calculation
        self.opf.iopt_obj = iopt_obj  # 'dev' Minimize control variable deviation, 'los' total losses
        self.opf.isWeightByCosts = 0 # Based on rated power/control variable range
        self.opf.iopt_pd = 1
        self.opf.iopt_qd = 1
        self.opf.iopt_trf = 1
        self.opf.iopt_sht = 0
        self.opf.iopt_brnch = 1
        self.opf.iopt_genP = 1
        self.opf.iopt_genQ = 1
        self.opf.iopt_bus = 1
        self.opf.iopt_add = 0
        self.opf.Execute()

        # Collect transformer data into a dictionary
        trf_data = {
            trf.loc_name: {"tap": trf.GetAttribute("c:nntap")}
            for trf in self.trfObj if not trf.GetAttribute("e:outserv")
        }

        # Collect generator data into a dictionary
        gen_data = {
            gen.loc_name: {
                "parallel": gen.ngnum,
                "P": gen.GetAttribute("m:P:bus1"),
                "Q": gen.GetAttribute("m:Q:bus1"),
                "pf": gen.GetAttribute("m:cosphi:bus1"),
                "V": gen.GetAttribute("m:u1:bus1"),
            }
            for gen in self.genObj if not gen.GetAttribute("e:outserv")
        }

        # Apply transformer tap settings
        for trf in self.trfObj:
            if trf.loc_name in trf_data:
                trf.nntap = int(trf_data[trf.loc_name]["tap"])

        # Apply generator settings
        for gen in self.genObj:
            if gen.loc_name in gen_data:
                gen_info = gen_data[gen.loc_name]
                gen.iv_mode = 1
                gen.mode_inp = 'DEF'
                gen.pgini = float(gen_info["P"] / gen_info["parallel"])
                gen.usetp = float(gen_info["V"])
                gen.cosgini = float(gen_info["pf"])
                gen.pf_recap = 0 if gen_info["Q"] > 0 else 1

    def getDispatch(self):
        gen = pd.DataFrame({
            "name": [obj.loc_name for obj in self.genObj],
            "P": [obj.GetAttribute("m:P:bus1") for obj in self.genObj],
            "Q": [obj.GetAttribute("m:Q:bus1") for obj in self.genObj],
            "pf": [obj.GetAttribute("m:cosphi:bus1") for obj in self.genObj],
            "V": [obj.GetAttribute("m:u1:bus1") for obj in self.genObj],
            "parallel": [obj.ngnum for obj in self.genObj],
        })
        return gen

    # ==========================================================================

    def short_circuit_setup(self, fault_location, fault_line, fault_duration):
        events = self.app.GetFromStudyCase('IntEvt')
        event_list = events.GetContents()
        
        # Clean up old events
        if event_list:
            for event in event_list:
                event.Delete()
        
        # Create new fault events
        events.CreateObject('EvtShc', 'SC_on_Line')
        shc_event = events.GetContents('SC_on_Line.EvtShc')[0]
        shc_event.i_shc = 0
        
        events.CreateObject('EvtSwitch', 'Trip_Line')
        switch_event = events.GetContents('Trip_Line.EvtSwitch')[0]
        
        # Configure fault line settings
        fault_line.ishclne = 1
        fault_line.fshcloc = fault_location
        shc_event.time = 0
        shc_event.p_target = fault_line
        switch_event.time = shc_event.time + fault_duration
        switch_event.p_target = fault_line

    def simulation_rms(self, monitored_variables, t_start=-100, t_step=10, t_stop=30):
        self.res = self.app.GetFromStudyCase('*.ElmRes')
        
        # Add monitored variables to result object
        for elm_name, var_names in monitored_variables.items():
            elements = self.app.GetCalcRelevantObjects(elm_name)
            for element in elements:
                self.res.AddVars(element, *var_names)

        # Setup simulation parameters
        self.inc = self.app.GetFromStudyCase('ComInc')
        self.sim = self.app.GetFromStudyCase('ComSim')
        self.ldf = self.app.GetFromStudyCase('ComLdf')
        
        # Set simulation options
        self.ldf.iopt_plim = 1
        self.ldf.iopt_lim = 1
        self.inc.iopt_sim = 'rms'
        self.inc.iopt_adapt = 1
        self.inc.tstart = t_start
        self.inc.dtgrd = t_step
        self.inc.dtgrd_max = t_step
        self.inc.iopt_sync = 1
        self.inc.syncperiod = t_step
        self.inc.ciopt_sample = 2
        self.sim.tstop = t_stop
        
        # Execute the simulation
        self.inc.Execute()
        self.sim.Execute()

    def save_load_and_gen(self, load_level, path):
        op_scen = f'{int(round(load_level*100, 0))} persen beban'
        
        wb = Workbook()
        wb.remove(wb.active)

        self._save_load_data(wb)
        self._save_gen_data(wb)
        self._save_transformer_data(wb)
        
        wb.save(os.path.join(path, f'Kondisi {op_scen}.xlsx'))

    def _save_load_data(self, wb):
        ws = wb.create_sheet('Beban')
        ws.cell(row=1, column=1).value = "Nama"
        ws.cell(row=1, column=2).value = "Daya aktif (MW)"
        ws.cell(row=1, column=3).value = "Daya reaktif (Mvar)"
        
        load_objects = self.app.GetCalcRelevantObjects("*.ElmLod")
        for load in load_objects:
            ws.append([load.loc_name, load.GetAttribute("m:P:bus1"), load.GetAttribute("m:Q:bus1")])

    def _save_gen_data(self, wb):
        ws = wb.create_sheet('Generator')
        ws.cell(row=1, column=1).value = "Nama"
        ws.cell(row=1, column=2).value = "Jumlah unit"
        ws.cell(row=1, column=3).value = "Tegangan (pu)"
        ws.cell(row=1, column=4).value = "Faktor daya"
        ws.cell(row=1, column=5).value = "Daya aktif (MW)"
        ws.cell(row=1, column=6).value = "Daya reaktif (MW)"
        
        gen_objects = self.app.GetCalcRelevantObjects("*.ElmSym")
        for gen in gen_objects:
            if not gen.GetAttribute("e:outserv"):
                ws.append([
                    gen.loc_name, gen.ngnum, 
                    gen.GetAttribute("m:u1:bus1"), gen.GetAttribute("m:cosphi:bus1"),
                    gen.GetAttribute("m:P:bus1"), gen.GetAttribute("m:Q:bus1")
                ])

    def _save_transformer_data(self, wb):
        ws = wb.create_sheet('Trafo')
        ws.cell(row=1, column=1).value = "Nama"
        ws.cell(row=1, column=2).value = "Posisi tap"
        
        trf_objects = self.app.GetCalcRelevantObjects("*.ElmTr2")
        for trf in trf_objects:
            if not trf.GetAttribute("e:outserv"):
                ws.append([trf.loc_name, trf.GetAttribute("c:nntap")])

    # Additional methods follow similar refactoring patterns...
